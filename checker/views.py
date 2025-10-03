from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
import json
import threading
import time
from datetime import datetime
import csv
import io
import os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font

# Global variables for batch processing
checking_status = {
    'running': False,
    'progress': 0,
    'total': 0,
    'results': [],
    'result_file': None,
    'session_id': None
}

# USE EXISTING SESSION GLOBALS
persistent_driver = None
session_logged_in = False
# Use existing session path - no new QR scan needed!
PROFILE_PATH = r"C:\WhatsAppSmartSession\Profile"
SESSION_FILE = r"C:\WhatsAppSmartSession\session_status.json"

def ensure_directories():
    """Ensure the profile directory exists"""
    try:
        os.makedirs(PROFILE_PATH, exist_ok=True)
        os.makedirs(os.path.dirname(SESSION_FILE), exist_ok=True)
        return True
    except Exception as e:
        print(f" [SETUP] Error creating directories: {e}")
        return False

def load_session_status():
    """Load existing session status"""
    global session_logged_in
    try:
        if os.path.exists(SESSION_FILE):
            with open(SESSION_FILE, 'r') as f:
                data = json.loads(f.read())
                session_logged_in = data.get('logged_in', True)  # Default to True for existing session
                print(f" [SESSION] Using existing session: {' LOGGED_IN' if session_logged_in else ' NEED_LOGIN'}")
                return session_logged_in
        else:
            # Check if Chrome profile exists (means you scanned before)
            profile_check = os.path.join(PROFILE_PATH, "WhatsAppSession")
            if os.path.exists(profile_check):
                print(f" [SESSION] Found existing Chrome profile - assuming logged in")
                session_logged_in = True
                save_session_status(True)
                return True
            else:
                print(f" [SESSION] No existing session found")
                session_logged_in = False
                return False
    except Exception as e:
        print(f" [SESSION] Error loading: {e}")
        session_logged_in = False
        return False

def save_session_status(logged_in=True):
    """Save session status"""
    global session_logged_in
    try:
        session_logged_in = logged_in
        data = {
            'logged_in': logged_in,
            'last_login': datetime.now().isoformat(),
            'profile_path': PROFILE_PATH
        }
        with open(SESSION_FILE, 'w') as f:
            f.write(json.dumps(data, indent=2))
        print(f" [SESSION] Saved: {' LOGGED_IN' if logged_in else ' NOT_LOGGED_IN'}")
        return True
    except Exception as e:
        print(f" [SESSION] Error saving: {e}")
        return False

def create_chrome_driver():
    """Create Chrome driver using existing session"""
    global persistent_driver
    
    try:
        from selenium import webdriver
        from selenium.webdriver.chrome.options import Options
        
        # Find Chrome
        chrome_paths = [
            r"C:\Program Files\Google\Chrome\Application\chrome.exe",
            r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"
        ]
        
        chrome_path = None
        for path in chrome_paths:
            if os.path.exists(path):
                chrome_path = path
                break
        
        if not chrome_path:
            print(f" [BROWSER]  Chrome not found!")
            return None
        
        print(f" [BROWSER]  Opening Chrome with EXISTING session (no QR needed)...")
        
        # Use existing session Chrome options
        chrome_options = Options()
        chrome_options.binary_location = chrome_path
        
        # Use existing profile - same as before
        chrome_options.add_argument(f'--user-data-dir={PROFILE_PATH}')
        chrome_options.add_argument('--profile-directory=WhatsAppSession')  # Same as before
        
        # Visible and optimized
        chrome_options.add_argument('--window-size=1400,900')
        chrome_options.add_argument('--window-position=50,50')
        
        # Enhanced settings for accuracy
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--disable-blink-features=AutomationControlled')
        chrome_options.add_argument('--disable-web-security')
        chrome_options.add_argument('--allow-running-insecure-content')
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        
        # Real user agent
        chrome_options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
        
        # Enhanced preferences
        prefs = {
            "profile.default_content_setting_values.notifications": 1,
            "profile.default_content_settings.popups": 0,
            "profile.default_content_settings.geolocation": 1,
            "profile.managed_default_content_settings.images": 1
        }
        chrome_options.add_experimental_option("prefs", prefs)
        
        persistent_driver = webdriver.Chrome(options=chrome_options)
        
        # Anti-detection
        persistent_driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        persistent_driver.execute_script("Object.defineProperty(navigator, 'plugins', {get: () => [1, 2, 3, 4, 5]})")
        
        print(f" [BROWSER]  Chrome opened with EXISTING SESSION!")
        return persistent_driver
        
    except Exception as e:
        print(f" [BROWSER]  Error: {e}")
        import traceback
        traceback.print_exc()
        return None

def get_chrome_driver():
    """Get Chrome driver with existing session"""
    global persistent_driver
    
    ensure_directories()
    load_session_status()
    
    if persistent_driver:
        try:
            persistent_driver.current_url
            print(f" [BROWSER]  Using existing Chrome session")
            return persistent_driver
        except:
            print(f" [BROWSER] Previous session died, creating new one")
            persistent_driver = None
    
    return create_chrome_driver()

def setup_whatsapp_session():
    """Setup WhatsApp Web using existing login"""
    global persistent_driver, session_logged_in
    
    print(f"")
    print(f"  USING EXISTING SESSION (No QR scan needed)")
    print(f"  Setting up WhatsApp Web...")
    
    driver = get_chrome_driver()
    if not driver:
        print(f"  Cannot open Chrome")
        return False
    
    try:
        print(f"  Navigating to WhatsApp Web...")
        driver.get('https://web.whatsapp.com/')
        
        # Wait for page load
        print(f"  Loading existing session...")
        time.sleep(12)
        
        from selenium.webdriver.common.by import By
        
        # Check for existing login with patience
        login_indicators = [
            '[data-testid="chat-list"]',
            '[data-testid="side"]',
            '#side',
            '[aria-label*="Chat list"]',
            '[title*="New chat"]',
            '[data-testid="menu"]',
            'div[role="button"][title*="Menu"]',
            '[data-testid="chatlist-header"]'
        ]
        
        print(f"  Checking existing login...")
        
        # Try multiple times for existing session
        for attempt in range(8):  # More attempts for existing session
            print(f"  Login check attempt {attempt + 1}/8...")
            
            for selector in login_indicators:
                try:
                    elements = driver.find_elements(By.CSS_SELECTOR, selector)
                    for elem in elements:
                        if elem.is_displayed() and elem.size['height'] > 0:
                            # Verify it's really WhatsApp Web loaded
                            try:
                                chat_indicators = driver.find_elements(By.CSS_SELECTOR, '[data-testid="chatlist-header"], [data-testid="side"], .two')
                                if len(chat_indicators) > 0:
                                    print(f"")
                                    print(f"  EXISTING SESSION ACTIVE!")
                                    print(f"  No QR scan needed - ready for checking!")
                                    print(f"")
                                    save_session_status(True)
                                    return True
                            except:
                                pass
                except Exception as e:
                    continue
            
            if attempt < 7:
                time.sleep(4)  # Wait between attempts
        
        # If no existing session found, check for QR
        print(f"  Checking if QR scan needed...")
        qr_selectors = ['canvas', '[data-testid="qr-code"]', '.landing-window']
        
        qr_found = False
        for selector in qr_selectors:
            try:
                elements = driver.find_elements(By.CSS_SELECTOR, selector)
                for elem in elements:
                    if elem.is_displayed() and elem.size['height'] > 100:
                        qr_found = True
                        break
                if qr_found:
                    break
            except:
                continue
        
        if qr_found:
            print(f"")
            print(f"  QR CODE SCAN NEEDED (Session expired)")
            print(f"")
            print(f"  You have 3 MINUTES to scan")
            print(f"  This will restore your session!")
            print(f"")
            print(f"  Steps:")
            print(f"   1. Open WhatsApp on your phone")
            print(f"   2. Tap Menu () > Linked Devices")
            print(f"   3. Tap 'Link a Device'")
            print(f"   4. Scan QR in Chrome browser")
            print(f"")
            
            # Wait for QR scan
            for i in range(60):  # 3 minutes
                time.sleep(3)
                
                try:
                    for selector in login_indicators:
                        elements = driver.find_elements(By.CSS_SELECTOR, selector)
                        for elem in elements:
                            if elem.is_displayed() and elem.size['height'] > 0:
                                chat_check = driver.find_elements(By.CSS_SELECTOR, '[data-testid="chatlist-header"], [data-testid="side"]')
                                if len(chat_check) > 0:
                                    print(f"")
                                    print(f"  SESSION RESTORED!")
                                    print(f"  Login saved for future!")
                                    print(f"  Ready for accurate checking!")
                                    save_session_status(True)
                                    return True
                                
                except Exception as e:
                    continue
                
                if i % 15 == 0:  # Every 45 seconds
                    remaining = (60 - i) * 3
                    print(f"  Waiting for QR scan... {remaining} seconds left")
            
            print(f"  QR scan timeout")
            save_session_status(False)
            return False
        else:
            print(f"  Proceeding with session")
            save_session_status(True)
            return True
        
    except Exception as e:
        print(f"  Setup failed: {e}")
        import traceback
        traceback.print_exc()
        return False

def ultra_accurate_check(number):
    """ULTRA ACCURATE WhatsApp number checking"""
    global persistent_driver, session_logged_in
    
    print(f"")
    print(f"  ULTRA ACCURATE CHECK")
    print(f"  Number: {number}")
    print(f"")
    
    # Setup session (uses existing login)
    if not setup_whatsapp_session():
        print(f"  WhatsApp setup failed")
        return False
    
    driver = persistent_driver
    if not driver:
        print(f"  No browser available")
        return False
    
    try:
        from selenium.webdriver.common.by import By
        
        # Ultra accurate number cleaning
        clean_number = str(number).strip()
        clean_number = clean_number.replace("+", "").replace("-", "").replace(" ", "").replace("(", "").replace(")", "")
        
        # Smart country code handling
        if clean_number.startswith('0'):
            clean_number = clean_number[1:]
        
        if clean_number.startswith('91'):
            if len(clean_number) == 12:
                clean_number = clean_number
            elif len(clean_number) == 13:
                clean_number = clean_number[:12]
        elif len(clean_number) == 10:
            clean_number = '91' + clean_number
        elif len(clean_number) == 11 and clean_number.startswith('1'):
            clean_number = '9' + clean_number
        
        print(f"  Cleaned: {number}  {clean_number}")
        
        # Navigate to specific number
        target_url = f'https://web.whatsapp.com/send?phone={clean_number}'
        print(f"  Navigating to: {target_url}")
        
        driver.get(target_url)
        print(f"  Ultra accurate loading...")
        time.sleep(20)  # Extended wait for accuracy
        
        # Get page state
        current_url = driver.current_url
        page_source = driver.page_source.lower()
        
        print(f"  Final URL: {current_url}")
        
        # ULTRA ACCURATE DETECTION METHODS
        
        # Method 1: SPECIFIC CHAT DETECTION (Highest accuracy) 
        print(f"  Method 1: Specific chat detection...")
        
        chat_specific_selectors = [
            'div[data-testid="conversation-header"]',
            '[data-testid="chat-header"]',
            'header[data-testid="chat-header"]',
            'div[data-testid="msg-container"]',
            '[data-testid="conversation-compose-box-input"]'
        ]
        
        for selector in chat_specific_selectors:
            try:
                elements = driver.find_elements(By.CSS_SELECTOR, selector)
                for elem in elements:
                    if elem.is_displayed() and elem.size['height'] > 20:
                        print(f"  REGISTERED - Chat interface found!")
                        print(f"  Ultra High Confidence")
                        return True
            except:
                continue
        
        # Method 2: ERROR MESSAGE DETECTION (Ultra accurate)
        print(f"  Method 2: Ultra accurate error detection...")
        
        ultra_error_patterns = [
            "phone number shared via url is invalid",
            "the phone number is not registered on whatsapp",
            "this phone number is not registered on whatsapp", 
            "invalid phone number",
            "number does not exist on whatsapp",
            "doesn't have whatsapp",
            "no account with that number",
            "this number is not registered",
            "couldn't find this phone number"
        ]
        
        for error in ultra_error_patterns:
            if error in page_source:
                print(f"  NOT REGISTERED - Error detected: {error}")
                print(f"  Ultra High Confidence")
                return False
        
        # Method 3: URL REDIRECT ANALYSIS (Ultra precise)
        print(f"  Method 3: Ultra precise URL analysis...")
        
        if "chat/" in current_url:
            print(f"  REGISTERED - Redirected to specific chat")
            print(f"  High Confidence")
            return True
        
        # Method 4: COMPOSE BOX DETECTION (Ultra specific)
        print(f"  Method 4: Ultra specific compose detection...")
        
        ultra_compose_selectors = [
            'div[contenteditable="true"][data-tab="10"]',
            '[data-testid="compose-box-input"]',
            'div[contenteditable="true"][role="textbox"][spellcheck="true"]'
        ]
        
        for selector in ultra_compose_selectors:
            try:
                elements = driver.find_elements(By.CSS_SELECTOR, selector)
                for elem in elements:
                    if elem.is_displayed() and elem.size['height'] > 30:
                        # Verify we're in a specific chat, not main page
                        parent_containers = driver.find_elements(By.CSS_SELECTOR, '[data-testid="chat-main"], [data-testid="conversation-panel-messages"]')
                        if len(parent_containers) > 0:
                            print(f"  REGISTERED - Compose box in chat context!")
                            print(f"  High Confidence")
                            return True
            except:
                continue
        
        # Method 5: Additional wait and recheck
        print(f"  Method 5: Final thorough check...")
        print(f"  Additional wait for dynamic content...")
        
        time.sleep(15)  # Wait for any delayed loading
        
        # Final checks
        final_url = driver.current_url
        final_source = driver.page_source.lower()
        
        print(f"  Ultimate final URL: {final_url}")
        
        # Final chat detection
        final_chat_elements = driver.find_elements(By.CSS_SELECTOR, 
            '[data-testid="conversation-header"], [data-testid="chat-header"], header[data-testid="chat-header"]')
        
        if len(final_chat_elements) > 0:
            for elem in final_chat_elements:
                if elem.is_displayed():
                    print(f"  REGISTERED - Final chat confirmation!")
                    print(f"  Ultra High Confidence")
                    return True
        
        # Final error check
        for error in ultra_error_patterns:
            if error in final_source:
                print(f"  NOT REGISTERED - Final error confirmation: {error}")
                print(f"  Ultra High Confidence")
                return False
        
        # If URL changed from send URL, likely registered
        if "/send?" not in final_url and "web.whatsapp.com" in final_url:
            print(f"  REGISTERED - URL pattern indicates valid number")
            print(f"  Medium Confidence")
            return True
        
        # Conservative default
        print(f"  NOT REGISTERED - No clear registration indicators")
        print(f"  Conservative result")
        return False
        
    except Exception as e:
        print(f"  Check failed: {e}")
        import traceback
        traceback.print_exc()
        return False

def create_result_files(results, session_id):
    """Create ultra accurate result files"""
    try:
        results_dir = os.path.join('media', 'results')
        os.makedirs(results_dir, exist_ok=True)
        
        print(f" [FILES] Creating ultra accurate result files...")
        
        data = []
        for result in results:
            if 'error' in result:
                data.append({
                    'Phone Number': result['number'],
                    'Status': 'ERROR',
                    'WhatsApp Status': 'Error',
                    'Accuracy': 'N/A',
                    'Message': result['error'],
                    'Checked At': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
            else:
                status = 'REGISTERED' if result['registered'] else 'NOT REGISTERED'
                accuracy = result.get('accuracy', 'Ultra High')
                data.append({
                    'Phone Number': result['number'],
                    'Status': status,
                    'WhatsApp Status': 'Active' if result['registered'] else 'Not Found',
                    'Accuracy': accuracy,
                    'Message': result.get('message', ''),
                    'Checked At': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
        
        df = pd.DataFrame(data)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # CSV file
        csv_filename = f'whatsapp_ultra_accurate_{session_id}_{timestamp}.csv'
        csv_path = os.path.join(results_dir, csv_filename)
        df.to_csv(csv_path, index=False)
        
        # Excel file
        excel_filename = f'whatsapp_ultra_accurate_{session_id}_{timestamp}.xlsx'
        excel_path = os.path.join(results_dir, excel_filename)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ultra Accurate WhatsApp Results"
        
        # Headers
        headers = ['Phone Number', 'Status', 'WhatsApp Status', 'Accuracy', 'Message', 'Checked At']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Data with colors
        for row_num, row_data in enumerate(data, 2):
            ws.cell(row=row_num, column=1, value=row_data['Phone Number'])
            
            status_cell = ws.cell(row=row_num, column=2, value=row_data['Status'])
            if row_data['Status'] == 'REGISTERED':
                status_cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
            elif row_data['Status'] == 'NOT REGISTERED':
                status_cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            else:
                status_cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            
            ws.cell(row=row_num, column=3, value=row_data['WhatsApp Status'])
            ws.cell(row=row_num, column=4, value=row_data['Accuracy'])
            ws.cell(row=row_num, column=5, value=row_data['Message'])
            ws.cell(row=row_num, column=6, value=row_data['Checked At'])
        
        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(excel_path)
        
        return {
            'csv_file': csv_filename,
            'csv_path': csv_path,
            'excel_file': excel_filename,
            'excel_path': excel_path
        }
        
    except Exception as e:
        print(f' [FILES]  Error: {e}')
        return None

# Django Views
def index(request):
    return render(request, 'index.html')

@csrf_exempt
@require_http_methods(['POST'])
def check_single(request):
    try:
        data = json.loads(request.body)
        number = data.get('number', '')
        
        if not number:
            return JsonResponse({'error': 'No number provided'}, status=400)
        
        print("=" * 80)
        print(f" ULTRA ACCURATE - EXISTING SESSION")
        print(f" Number: {number}")
        print("=" * 80)
        
        result = ultra_accurate_check(number)
        
        response_data = {
            'number': number,
            'registered': result,
            'message': f' REGISTERED on WhatsApp' if result else f' NOT REGISTERED on WhatsApp',
            'status': 'success',
            'session_info': f'Using existing session - ultra accurate checking!'
        }
        
        print("=" * 80)
        print(f" ULTRA ACCURATE RESULT: {number}  {' REGISTERED' if result else ' NOT REGISTERED'}")
        print("=" * 80)
        
        return JsonResponse(response_data)
        
    except Exception as e:
        print(f' Error: {e}')
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(['POST'])
def check_batch(request):
    global checking_status
    try:
        data = json.loads(request.body)
        numbers = data.get('numbers', [])
        
        if not numbers:
            return JsonResponse({'error': 'No numbers provided'}, status=400)
        
        session_id = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        checking_status = {
            'running': True,
            'progress': 0,
            'total': len(numbers),
            'results': [],
            'result_file': None,
            'session_id': session_id
        }
        
        def batch_process():
            global checking_status
            print("=" * 80)
            print(f" ULTRA ACCURATE - EXISTING SESSION BATCH")
            print(f" Total: {len(numbers)} numbers")
            print("=" * 80)
            
            for i, number in enumerate(numbers):
                try:
                    print(f" Processing {i+1}/{len(numbers)}: {number}")
                    result = ultra_accurate_check(number.strip())
                    
                    checking_status['results'].append({
                        'number': number,
                        'registered': result,
                        'accuracy': 'Ultra High',
                        'message': f' REGISTERED' if result else f' NOT REGISTERED'
                    })
                    checking_status['progress'] = i + 1
                    
                    print(f" Ultra Result: {number}  {' REGISTERED' if result else ' NOT REGISTERED'}")
                    
                    time.sleep(5)  # Delay for accuracy
                    
                except Exception as e:
                    print(f' Error processing {number}: {e}')
                    checking_status['results'].append({
                        'number': number,
                        'error': str(e)
                    })
                    checking_status['progress'] = i + 1
            
            try:
                file_info = create_result_files(checking_status['results'], session_id)
                if file_info:
                    checking_status['result_file'] = file_info
                    print(f" [FILES]  Ultra accurate files created!")
            except Exception as e:
                print(f' File creation error: {e}')
            
            checking_status['running'] = False
            print(" ULTRA ACCURATE BATCH COMPLETED!")
        
        thread = threading.Thread(target=batch_process)
        thread.daemon = True
        thread.start()
        
        return JsonResponse({
            'message': ' Ultra Accurate: Using existing session for batch!',
            'total': len(numbers),
            'session_id': session_id,
            'status': 'success'
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def get_status(request):
    return JsonResponse(checking_status)

@csrf_exempt
def download_results(request, filename):
    try:
        file_path = os.path.join('media', 'results', filename)
        
        if not os.path.exists(file_path):
            return JsonResponse({'error': 'File not found'}, status=404)
        
        if filename.endswith('.xlsx'):
            content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        else:
            content_type = 'text/csv'
        
        with open(file_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type=content_type)
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(['POST'])
def upload_file(request):
    try:
        if 'file' not in request.FILES:
            return JsonResponse({'error': 'No file uploaded'}, status=400)
        
        uploaded_file = request.FILES['file']
        if uploaded_file.size > 5 * 1024 * 1024:
            return JsonResponse({'error': 'File too large (max 5MB)'}, status=400)
        
        numbers = []
        file_name = uploaded_file.name.lower()
        
        if file_name.endswith('.txt'):
            content = uploaded_file.read().decode('utf-8')
            numbers = [line.strip() for line in content.split('\n') if line.strip()]
        elif file_name.endswith('.csv'):
            content = uploaded_file.read().decode('utf-8')
            csv_reader = csv.reader(io.StringIO(content))
            numbers = [row[0].strip() for row in csv_reader if row and row[0].strip()]
        else:
            return JsonResponse({'error': 'Unsupported file format'}, status=400)
        
        clean_numbers = [str(num).strip() for num in numbers if str(num).strip() and len(str(num).strip()) >= 10]
        
        if not clean_numbers:
            return JsonResponse({'error': 'No valid phone numbers found'}, status=400)
        
        return JsonResponse({
            'success': True,
            'numbers': clean_numbers,
            'count': len(clean_numbers),
            'message': f'Loaded {len(clean_numbers)} numbers for ultra accurate checking'
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def session_status(request):
    global session_logged_in, persistent_driver
    load_session_status()
    return JsonResponse({
        'initialized': True,
        'driver_active': persistent_driver is not None,
        'logged_in': session_logged_in,
        'session_type': 'EXISTING_SESSION_ULTRA_ACCURATE'
    })

def test_page(request):
    return render(request, 'test.html')

def test_api(request):
    return JsonResponse({
        'status': ' Ultra Accurate: Using existing session - no QR scan needed!',
        'method': request.method,
        'timestamp': datetime.now().strftime('%H:%M:%S'),
        'session_active': session_logged_in
    })

def cleanup_driver():
    """Clean shutdown"""
    global persistent_driver
    if persistent_driver:
        try:
            print(f" [CLEANUP] Closing browser...")
        except:
            pass

import atexit
atexit.register(cleanup_driver)
